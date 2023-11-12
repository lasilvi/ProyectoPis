from .forms import *
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from .models import *
from openpyxl import load_workbook
from collections import Counter
from statistics import mean
from django.http import JsonResponse
from django.db.models import Count

def login_view(request):
    if request.method == 'POST':
        form = UserLoginForm(request=request, data=request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('importar_datos')
            else:
                return render(request=request, template_name="registro/login.html", context={'form': form})
        else:
            for key, error in list(form.errors.items()):
                messages.error(request, error)
    form = UserLoginForm()
    return render(request=request, template_name="registro/login.html", context={'form': form})

def logout_view(request):
    logout(request)
    return redirect('login') 

@login_required
def viewUCE(request):
    Equipo = Equipos.objects.filter(servicio="UCE")

    # Obtener sugerencias desde la base de datos
    sugerencias_ubicacion = list(obtener_sugerencias("UCE",'ubicacion'))
    sugerencias_nombre = list(obtener_sugerencias("UCE",'nombre'))
    sugerencias_bio = list(obtener_sugerencias("UCE",'bio'))


    conteo_equipos = contar_equipos_por_nombres("UCE")
    
    promedio_cumplimiento = calcular_porcentaje_cumplimiento_total("UCE")

    consultorios = Equipos.objects.filter(servicio="UCE").values('ubicacion').distinct()
    # Filtra en el lado de Python para contar solo las ubicaciones numéricas
    total_consultorios = sum(1 for consultorio in consultorios if str(consultorio['ubicacion']).isdigit())


    contexto = {
        "promedio_cumplimiento":promedio_cumplimiento,
        "equipos": Equipo,
        "conteo_equipos":conteo_equipos,
        "sugerencias_ubicacion": sugerencias_ubicacion,
        "sugerencias_nombre": sugerencias_nombre,
        "sugerencias_bio": sugerencias_bio,
        "total_consultorios": total_consultorios
        # Otros datos del contexto que desees pasar a la plantilla
    }
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == "filtrar":
            # El usuario hizo una solicitud de filtrado
            servicio = "UCE"
            ubicacion = request.POST.get('ubicacion')
            bio = request.POST.get('bio')
            nombre = request.POST.get('nombre')
        
            print(ubicacion)
            print(bio)

            equipos_filtrados = buscar_equipos(servicio, ubicacion, bio,nombre)

            return render(request, 'appseguimiento/UCE.html', {
                "promedio_cumplimiento":promedio_cumplimiento,
                "conteo_equipos":conteo_equipos,
                'equipos_filtrados': equipos_filtrados,
                'servicio': servicio,
                'ubicacion': ubicacion,
                'bio': bio,
                'nombre': nombre,
                "sugerencias_ubicacion": sugerencias_ubicacion,
                "sugerencias_nombre": sugerencias_nombre,
                "sugerencias_bio": sugerencias_bio,
                "total_consultorios": total_consultorios
            })
        if action == "verConsultorio":
            ubicacionconsultorio = request.POST.get('ubicacionconsultorio')
            equiposconsultorio = Equipos.objects.filter(servicio="UCE",ubicacion=ubicacionconsultorio).values("nombre").distinct()
            return render(request, 'appseguimiento/UCE.html', {
                "promedio_cumplimiento":promedio_cumplimiento,
                "equipos": Equipo,
                "conteo_equipos":conteo_equipos,
                "sugerencias_ubicacion": sugerencias_ubicacion,
                "sugerencias_nombre": sugerencias_nombre,
                "sugerencias_bio": sugerencias_bio,
                "total_consultorios": total_consultorios,
                "equiposconsultorio":equiposconsultorio
            })
           
    
    return render(request, 'appseguimiento/UCE.html', contexto)


def obtener_sugerencias(servicio,filtro):
    # Aquí deberías consultar tu base de datos para obtener las sugerencias según el filtro
    # Supongamos que tienes un modelo llamado Sugerencias con un campo 'valor' que contiene las sugerencias

    # values y annotate para obtener valores únicos y contar la cantidad de ocurrencias de cada uno en la columna correspondiente. Luego, extraemos los valores únicos con values_list.
    if filtro == 'ubicacion':
        sugerencias = Equipos.objects.filter(servicio=servicio).values('ubicacion').annotate(count=Count('ubicacion')).values_list('ubicacion', flat=True)
    elif filtro == 'nombre':
        sugerencias = Equipos.objects.filter(servicio=servicio).values('nombre').annotate(count=Count('nombre')).values_list('nombre', flat=True)
    elif filtro == 'bio':
        sugerencias = Equipos.objects.filter(servicio=servicio).values('bio').annotate(count=Count('bio')).values_list('bio', flat=True)
    else:
        sugerencias = []

    return list(sugerencias)

def buscar_equipos(servicio, ubicacion, bio, nombre):
    equipos = Equipos.objects.filter(servicio=servicio)

    if ubicacion:
        equipos = equipos.filter(ubicacion=ubicacion)

    if bio:
        equipos = equipos.filter(bio=bio)

    if nombre:
        equipos = equipos.filter(nombre=nombre)

    return equipos
    
def contar_equipos_por_nombres(servicio):
    sugerencias = Equipos.objects.filter(servicio=servicio).values('nombre').annotate(count=Count('nombre'))

    # Crear un diccionario a partir de los resultados
    diccionario_sugerencias = {item['nombre']: item['count'] for item in sugerencias}

    # Devuelve el resultado como un diccionario
    return diccionario_sugerencias


def calcular_porcentaje_cumplimiento_total(servicio):
    if not servicio:
        # Handle the case where the data set is empty
        return None
    
    equipos = Equipos.objects.filter(servicio=servicio)
    porcentajes_cumplimiento = []

    for equipo in equipos:
        criterios = [
            equipo.etiqueta_caja,
            equipo.factura_contrato,
            equipo.resolucion_invima,
            equipo.carta_garantia,
            equipo.acta_entrega,
            equipo.declaracion_importacion,
            equipo.cronograma_mantenimiento,
            equipo.cronograma_calibracion,
            equipo.reportes_mantenimiento,
            equipo.reportes_calibracion,
            equipo.guia_rapida,
            equipo.manual_usuario_espanol,
            equipo.manual_servicio_espanol,
            equipo.hoja_vida_calibracion,
            equipo.hoja_vida_mantenimiento,
            equipo.contrato_mantenimiento,
            equipo.contrato_calibracion,
        ]

        cumple = sum(1 for estado in criterios if estado == 'CUMPLE')
        no_aplica = sum(1 for estado in criterios if estado == 'NO APLICA')
        total_criterios = len(criterios)

        total_cumple_no_aplica = cumple + no_aplica

        if total_criterios == 0:
            porcentaje_cumplimiento = 0
        else:
            porcentaje_cumplimiento = (total_cumple_no_aplica / total_criterios) * 100

        porcentajes_cumplimiento.append(porcentaje_cumplimiento)
    
    if not porcentajes_cumplimiento:
        return 0

    # Calcula el promedio de los porcentajes de cumplimiento
    promedio_cumplimiento = mean(porcentajes_cumplimiento)
    promedio_cumplimiento = round(promedio_cumplimiento)

    return promedio_cumplimiento

def calcular_porcentaje_cumplimiento(equipo):
    criterios = {
        'etiqueta_caja': equipo[5],
        'factura_contrato': equipo[6],
        'resolucion_invima': equipo[7],
        'carta_garantia': equipo[8],
        'acta_entrega': equipo[9],
        'declaracion_importacion': equipo[10],
        'cronograma_mantenimiento': equipo[11],
        'cronograma_calibracion': equipo[12],
        'reportes_mantenimiento': equipo[13],
        'reportes_calibracion': equipo[14],
        'guia_rapida': equipo[15],
        'manual_usuario_espanol': equipo[16],
        'manual_servicio_espanol': equipo[17],
        'hoja_vida_calibracion': equipo[18],
        'hoja_vida_mantenimiento': equipo[19],
        'contrato_mantenimiento': equipo[20],
        'contrato_calibracion': equipo[21],
    }

    cumple = sum(1 for estado in criterios.values() if estado == 'CUMPLE')
    no_aplica = sum(1 for estado in criterios.values() if estado == 'NO APLICA')
    total_criterios = len(criterios)

    total_cumple_no_aplica = cumple + no_aplica

    if total_criterios == 0:
        return 0

    porcentaje_cumplimiento = round((total_cumple_no_aplica / total_criterios) * 100, 2)
    return porcentaje_cumplimiento

@login_required
def importar_datos(request):
    if request.method == 'POST':
        form = ImportarDatosForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = form.cleaned_data['archivo_excel']
            procesar_archivo_excel(archivo_excel)
            print("datos importados")
            return redirect('importar_datos') 
        else:
            print("no valido")# Redirige a la página de importación de datos
    else:
        form = ImportarDatosForm()
    return render(request, 'appseguimiento/index.html', {'form': form})

def procesar_archivo_excel(archivo_excel):
    wb = load_workbook(archivo_excel)
    hoja = wb.active

    # Extract the service from the first row of the Excel sheet
    servicio = hoja.cell(row=2, column=1).value
    print(servicio)

    # Check if there are existing Equipos for the specified service
    existing_equipos = Equipos.objects.filter(servicio=servicio)

    if existing_equipos.exists():
        # If existing Equipos are found for the service, delete only those
        existing_equipos.delete()


    for fila in hoja.iter_rows(min_row=2, values_only=True):
          # Supongamos que la primera columna contiene nombres y la segunda columnas contiene valores
        if any(fila[1:]): 
            # Supongamos que la primera columna contiene nombres y la segunda columnas contiene valores
            equipo = Equipos(
                servicio=fila[0], 
                bio=fila[1],
                nombre=fila[2],
                ubicacion=fila[3],
                sede=fila[4],
                etiqueta_caja=fila[5],
                factura_contrato=fila[6],
                resolucion_invima=fila[7],
                carta_garantia=fila[8],
                acta_entrega=fila[9],
                declaracion_importacion=fila[10],
                cronograma_mantenimiento=fila[11],
                cronograma_calibracion=fila[12],
                reportes_mantenimiento=fila[13],
                reportes_calibracion=fila[14],
                guia_rapida=fila[15],
                manual_usuario_espanol=fila[16],
                manual_servicio_espanol=fila[17],
                hoja_vida_calibracion=fila[18],
                hoja_vida_mantenimiento=fila[19],
                contrato_mantenimiento=fila[20],
                contrato_calibracion=fila[21],
                porcentajedecumplimiento = calcular_porcentaje_cumplimiento(fila),
            )
            equipo.save()
            